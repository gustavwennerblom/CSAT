import json
import logging
import mysql.connector
import os
from datetime import datetime
from sqlalchemy import create_engine, Table, MetaData, inspect
from sqlalchemy.sql import select, and_
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class CSSreply:
    sub_project_no = row[3],
    region = row[0],
    office = row[1],
    customer_name = row[2],
    pm_first_name = row[4],
    pm_last_name = row[5],
    date_answered = row[6],
    question_id = row[7],
    question_text = row[8],
    answer_numeric = row[9],
    answer_text = row[10])

    questions_and_answers={}

    def __init__(self,sub_project_no, question_id, question_text, **kwargs):
        self.sub_project_no=sub_project_no
        for key, value in kwargs.items():
            setattr(self, key, value)
            print(".")

    def add_answer(self):

    # def __repr__(self):
    #     return "CSSreply for subproject {0}".format(self.sub_project_no)

class CSATanalyzer:

    def __init__(self):
        # Create logger
        log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        logging.basicConfig(filename="./CSATstats.log", format=log_format, level=logging.INFO)

        # Read MySQL config and credentials and store in a dict
        with open(os.path.join(os.path.dirname(__file__), "creds_mysql.json")) as f:
            logging.info("Accessing database credentials")
            config = json.loads(f.readline())

        sqluser = config["sqluser"]
        sqlpassword = config["sqlpassword"]
        sqlhost = config["sqlhost"]
        database_name = config["database"]

        self.conn = mysql.connector.connect(user=sqluser, password=sqlpassword, host=sqlhost, database=database_name)
        self.cur = self.conn.cursor()

        self.db = create_engine("mysql+mysqlconnector://%s:%s@%s/%s" % (sqluser, sqlpassword, sqlhost, database_name))
        self.db.echo = False

        meta = MetaData()
        meta.reflect(bind=self.db)
        self.projects = Table('projects', meta, autoload=True)
        self.questions = Table('questions', meta, autoload=True)
        self.answers = Table('answers', meta, autoload=True)
        self.ratings = Table('ratings', meta, autoload=True)
        self.con = self.db.connect()
        logging.info("Connected to database")

        self.attribute_list = ["Office", "Client", "Project number", "Sent to PM",
                               "Sent by PM", "PM First Name", "PM Last Name", "Date of upload"]
        self.title_font = Font(bold=True)

    # Constructs a set of all offices in the database
    def build_office_set(self):
        stmt = select([self.projects.c.office])
        result = self.con.execute(stmt)
        office_set = set()
        for row in result:
            office_set.add(row[0])
        logging.info("Set of offices constructed from database")
        return office_set

    def build_region_set(self):
        stmt = select([self.projects.c.region])
        result = self.con.execute(stmt)
        region_set = set()
        for row in result:
            region_set.add(row[0])
        logging.info("Set of regions constructed from database")
        return region_set

    # Currently not in use
    def get_pending(self, office):
        stmt = select([
            self.projects.c.office,
            self.projects.c.projectName,
            self.projects.c.pmName,
            self.projects.c.pmLastName,
            self.projects.c.subProjectNo,
            self.projects.c.customerName,
            self.projects.c.externalContact,
            self.projects.c.contact
        ]).where(and_(
            self.projects.c.office == office,
            self.projects.c.pmSendStatus == "no",
            self.projects.c.adminSendStatus == "yes"
        ))
        result = self.con.execute(stmt)

        out = []
        for row in result:
            out.append(row)
        return out

    # Returns list of two ints, the first one representing total CSS's due to send by the office and the second one
    # the number of actually sent CSS's
    def count_pending(self, office, **kwargs):
        start_date = datetime(2017, 1, 1)
        try:
            assert isinstance(kwargs["start_date"], datetime)
            start_date = kwargs["start_date"]
        except TypeError:
            logging.info("No start date given, considering surveys triggered from {0} "
                         "and forward in time".format(start_date))

        stmt_total = select([self.projects.c.subProjectNo]).where(and_(
            self.projects.c.office == office,
            self.projects.c.dateUpload > start_date))

        stmt_pending = select([self.projects.c.subProjectNo]).where(and_(
            self.projects.c.office == office,
            self.projects.c.pmSendStatus == "no",
            self.projects.c.dateUpload > start_date))

        result = [len(self.con.execute(stmt_total).fetchall()), len(self.con.execute(stmt_pending).fetchall())]

        return result

    # test function
    def get_a_date(self):
        stmt = select([self.projects.c.dateUpload]).where(self.projects.c.office == "Shanghai")
        result = self.con.execute(stmt)
        one_date = result.fetchone()
        return 1

    # Returns a list of answered surveys for a given office:
    def get_answers_office(self, office, start_date):
        stmt = select([
            self.projects.c.office,
            self.projects.c.customerName,
            self.projects.c.subProjectNo,
            self.projects.c.pmName,
            self.projects.c.pmLastName,
            self.answers.c.dateAnswer,
            self.answers.c.questionId,
            self.questions.c.question,
            self.answers.c.answersNumeric,
            self.answers.c.answersText,
        ]).where(and_(
            self.answers.c.ratingId == self.ratings.c.ratingId,
            self.answers.c.questionId == self.questions.c.questionId,
            self.ratings.c.projectId == self.projects.c.projectId,
            self.projects.c.office == office,
            self.projects.c.dateUpload > start_date)
        ).order_by(
            self.projects.c.office,
            self.projects.c.subProjectNo,
            self.answers.c.answerId,
            self.answers.c.questionId
        )

        result = self.con.execute(stmt)
        logging.info("Queried database for surveys answered: office % s" % office)

        out = []
        for row in result:
            out.append(row)
        return out

    # Returns a list of answered surveys for a given region, provided a start date
    # for survey upload (datetime.datetime):
    def get_answers_region(self, region, start_date):
        stmt = select([
            self.projects.c.region,
            self.projects.c.office,
            self.projects.c.customerName,
            self.projects.c.subProjectNo,
            self.projects.c.pmName,
            self.projects.c.pmLastName,
            self.answers.c.dateAnswer,
            self.answers.c.questionId,
            self.questions.c.question,
            self.answers.c.answersNumeric,
            self.answers.c.answersText,
        ]).where(and_(
            self.answers.c.ratingId == self.ratings.c.ratingId,
            self.answers.c.questionId == self.questions.c.questionId,
            self.ratings.c.projectId == self.projects.c.projectId,
            self.projects.c.region == region,
            self.projects.c.dateUpload > start_date)
        ).order_by(
            self.projects.c.office,
            self.projects.c.subProjectNo,
            self.answers.c.answerId,
            self.answers.c.questionId
        )

        result = self.con.execute(stmt)
        logging.info("Queried database for surveys answered: region {0}".format(region))

        out = {}
        for row in result:
            sub_proj_no = row[3]
            if not sub_proj_no in out:
                response_new = CSSreply(sub_project_no=row[3],
                                    region=row[0],
                                    office=row[1],
                                    customer_name=row[2],
                                    pm_first_name=row[4],
                                    pm_last_name=row[5],
                                    date_answered=row[6],
                                    question_id=row[7],
                                    question_text=row[8],
                                    answer_numeric=row[9],
                                    answer_text=row[10])
                out[sub_proj_no]=response_new
            else:
                response_append = out[sub_proj_no]
                response_append.question_id
        return out

    # Returns list of pending surveys for a given region
    def get_pending_region(self, region, start_date):
        stmt = select([
            self.projects.c.office,
            self.projects.c.customerName,
            self.projects.c.subProjectNo,
            self.projects.c.adminSendStatus,
            self.projects.c.pmSendStatus,
            self.projects.c.pmName,
            self.projects.c.pmLastName,
            self.projects.c.dateUpload
        ]).where(and_(
            self.projects.c.region == region,
            self.projects.c.dateUpload > start_date
        )).order_by(
            self.projects.c.office,
            self.projects.c.subProjectNo,
            self.projects.c.adminSendStatus,
            self.projects.c.pmSendStatus)

        result = self.con.execute(stmt)
        logging.info("Queried database for surveys triggered")
        out = []
        for row in result:
            out.append(row)
        return out

    def print_all_pending_by_office(self, offices):
        counter = 0
        for office in offices[:2]:
            out = self.get_pending(office)
            for s in out:
                print(s)
                counter += 1
            print("---")
        print("%i surveys missing" % counter)

    def print_all_pending_by_region(self, regions, start_date):
        wb = Workbook()
        logging.info("Preparing results spreadsheet")
        for reg in regions:
            totals = {}
            ws = wb.create_sheet(title=reg)
            ws.cell(row=1, column=1).value = "Client Satisfaction Surveys for Region %s" % reg
            col = 1
            row = 3
            # Loop to write header row
            for header in self.attribute_list:
                ws.cell(row=row, column=col).value = header
                ws.cell(row=row, column=col).font = self.title_font
                col += 1
            row += 1
            col = 1

            # Collect list of pending surveys for the region
            missing_surveys = self.get_pending_region(reg, start_date)
            i = 0
            for survey in missing_surveys:
                totals[survey[0]] = []
                for val in survey:
                    try:
                        totals[survey[0]][i] += 1
                    except IndexError:
                        logging.info("Creating index for %s" % survey[0])
                        totals[survey[0]].append(0)
                    i += 1
                    ws.cell(row=row, column=col).value = val
                    col += 1
                    # print("¤ %s" % val)
                row += 1
                col = 1
                i = 0
                # print("---")

            row += 1
            ws.cell(row=row, column=col).value = "Totals by office"
            row += 1
            col = 1
            ws.cell(row=row, column=col).value = "Total sent by PM"
            col += 1
            ws.cell(row=row, column=col).value = "Total sent to client"
            col += 1
            ws.cell(row=row, column=col).value = "Total rejected"
            row += 1
            col = 1
            for key, values in totals.items():
                ws.cell(row=row, column=col).value = key
                col += 1
                ws.cell(row=row, column=col).value = values[0]
                col += 1
                ws.cell(row=row, column=col).value = values[1]
                col = 1
                row += 1

        # print("%i surveys missing" % counter)

        # Clean up autocreated blank sheets in workbook
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        logging.info("Cleaning up stuff...")
        # wb.remove_sheet(wb.get_sheet_by_name("Sheet1"))

        # Add a time/date stamp to filename and save
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d-%H%M")
        filename = "Missing CSS (regions) %s.xlsx" % timestamp
        wb.save(filename)
        logging.info("Excel file saved in program root with name %s" % filename)

    @staticmethod
    def alternating_fill(colorcode):
        color1 = PatternFill("solid", fgColor="00CCFF")
        color2 = PatternFill("solid", fgColor="FFFFFF")
        if colorcode == color1:
            return color2
        else:
            return color1

    # Outputs all answers for a set of offices into an Excel workbook (one tab per office),
    # given a specific start date
    def print_all_answers_by_office(self, answers, headers, start_date):
        wb = Workbook()
        logging.info("Preparing answers workbook")
        cellcolor = self.alternating_fill(PatternFill("solid", fgColor="FFFFFF"))

        total_answers = 0
        for office in answers:
            ws = wb.create_sheet(title=office)
            ws.cell(row=1, column=1).value = "Client Satisfaction Survey Answers for %s" % office
            col = 1
            row = 3
            # Loop to write header row
            for header in headers:
                ws.cell(row=row, column=col).value = header
                ws.cell(row=row, column=col).font = self.title_font
                col += 1
            row += 1
            col = 1

            answers_for_office = self.get_answers_office(office, start_date)
            total_answers += len(answers_for_office)
            logging.info("Got {0} answer lines for {1}, total found stands at {2}"
                         .format(len(answers_for_office), office, total_answers))

            for answer in answers_for_office:
                for data in answer:
                    ws.cell(row=row, column=col).value = data
                    col += 1
                # Change cell fill color between survey answers
                # cellcolor = alternating_fill(cellcolor)
                # ws.cell(row=row, column=col).fill = cellcolor
                # Put cursor on new row and first column
                row += 1
                col = 1

        # Clean up autocreated blank sheets in workbook
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        logging.info("Cleaning up stuff...")

        # Add a time/date stamp to filename and save
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d-%H%M")
        filename = "CSS answers (by office) %s.xlsx" % timestamp
        wb.save(filename)
        logging.info("Excel file saved in program root with name %s" % filename)

    # Generates an excel file with survey answers for all offices, one sheet per region, given a specific start date
    def print_all_answers_by_region(self, regions, headers, start_date):
        wb = Workbook()
        logging.info("Preparing answers workbook")
        cellcolor = self.alternating_fill(PatternFill("solid", fgColor="FFFFFF"))

        total_answers = 0
        for region in regions:
            ws = wb.create_sheet(title=region)
            ws.cell(row=1, column=1).value = "Client Satisfaction Survey Answers for %s" % region
            col = 1
            row = 3
            # Loop to write header row
            for header in headers:
                ws.cell(row=row, column=col).value = header
                ws.cell(row=row, column=col).font = self.title_font
                col += 1
            row += 1
            col = 1

            answers_for_region = self.get_answers_region(region, start_date)
            total_answers += len(answers_for_region)

            logging.info("Got {0} CSS replies  for {2}, total for this run stands at {1} lines"
                         .format(len(answers_for_region), total_answers, region))

            for reply in answers_for_region:
                for item in reply:
                    # assert isinstance(item, CSSreply)
                    ws.cell(row=row, column=1).value = item.office
                    ws.cell(row=row, column=2).value = item.customer_name
                    ws.cell(row=row, column=3).value = item.sub_project_no
                    ws.cell(row=row, column=4).value = item.pm_first_name
                    ws.cell(row=row, column=5).value = item.pm_last_name
                    ws.cell(row=row, column=6).value = item.date_answered
                    ws.cell(row=row, column=7).value = item.qu

                    col += 1
                row += 1
                col = 1

            response = CSSreply(sub_project_no=row[3],
                                region=row[0],
                                office=row[1],
                                customer_name=row[2],
                                pm_first_name=row[4],
                                pm_last_name=row[5],
                                date_answered=row[6],
                                question_id=row[7],
                                question_text=row[8],
                                answer_numeric=row[9],
                                answer_text=row[10])
            headers = ["Office", "Client", "Project no", "PM Name", "PM Last Name",
                       "Date answered", "Question number", "Question", "Answer (score)", "Answer (text/comment)"]

        # Clean up autocreated blank sheets in workbook
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        logging.info("Cleaning up stuff...")

        # Add a time/date stamp to filename and save
        timestamp = datetime.now().strftime("%Y%m%d-%H%M")
        filename = "CSS answers (by region) %s.xlsx" % timestamp
        wb.save(filename)
        logging.info("Excel file saved in program root with name %s" % filename)

    def inspect_table(self, tablename):
        inspector = inspect(self.db)
        for column in inspector.get_columns(tablename):
            print(column['name'])
        # stmt = select([projects])
        # out = con.execute(stmt)
        # print(out.fetchall())

    def get_fields(self):
        self.inspect_table("projects")

    def get_status_main(self):
        # inspect_table("projects")
        offices = list(self.build_office_set())
        regions = list(self.build_region_set())

        offices.sort()
        regions.sort()

        self.print_all_pending_by_region(regions)
        logging.info("All done, shutting down. Enjoy.")

    def get_answers_by_office_main(self):
        logging.info("Starting process to print workbook with survey answers by office")
        offices = list(self.build_office_set())
        offices.sort()
        headers = ["Office", "Client", "Project no", "PM Name", "PM Last Name",
                   "Date answered", "Question number", "Question", "Answer (score)", "Answer (text/comment)"]
        self.print_all_answers_by_office(offices, headers, datetime(2017, 1, 1))
        logging.info("All done, shutting down. Enjoy.")

    def get_answers_by_region_main(self):
        logging.info("Starting process to print workbook with survey answers by region")
        regions = list(self.build_region_set())
        regions.sort()
        headers = ["Region", "Office", "Client", "Project no", "PM Name", "PM Last Name",
                   "Date answered", "Question number", "Question", "Answer (score)", "Answer (text/comment)"]
        self.print_all_answers_by_region(regions, headers, datetime(2017, 1, 1))

    def get_pending_by_region_main(self, start_date):
        logging.info("Starting process to print workbook with send status by region")
        regions = list(self.build_region_set())
        self.print_all_pending_by_region(regions, start_date)


if __name__ == '__main__':
    ca = CSATanalyzer()

    print("Please select action:")
    print("[1] Create xlsx on pending surveys by Region since 2017/01/01")
    print("[2] Create xlsx on answers received by Region")
    print("[3] Create xlsx on answers received by Office")
    print("[4] Drop to shell")
    selection = int(input("..."))

    if selection == 1:
        ca.get_pending_by_region_main(datetime(2017, 1, 1))
    elif selection == 2:
        ca.get_answers_by_region_main()
    elif selection == 3:
        ca.get_answers_by_office_main()
    elif selection == 4:
        import code
        code.interact(local=locals())
    else:
        print("Bad input. Bye")

    print("Successful. Please check root directory")


    # Method to trigger printout of survey's pending to be sent
    # get_status_main()

    # Method to trigger collection of answers

    # offices = ca.build_office_set()
    # for office in offices:
    #     print("{1}: {2} pending, of total {0}".format(ca.count_pending(office, start_date=datetime(2017, 1, 1))[0], 
    #                                                   office, 
    #                                                   ca.count_pending(office, start_date=datetime(2017,1,1))[1]))

    # ca.get_answers_by_office_main()
    #
