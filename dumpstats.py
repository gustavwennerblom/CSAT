import json
import logging
import mysql.connector
from sqlalchemy import create_engine, Table, MetaData, inspect
from sqlalchemy.sql import select, and_, or_
from sqlalchemy.orm import sessionmaker
from openpyxl import Workbook
from openpyxl.styles import Font

# Create logger
FORMAT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
logging.basicConfig(filename="CSATstats.log", format=FORMAT, level=logging.INFO)

# Read MySQL config and credentials and store in a dict
with open("creds_mysql.json") as f:
    logging.info("Accessing database credentials")
    config = json.loads(f.readline())

sqluser = config["sqluser"]
sqlpassword = config["sqlpassword"]
sqlhost = config["sqlhost"]
database_name = config["database"]

conn = mysql.connector.connect(user=sqluser, password=sqlpassword, host=sqlhost, database=database_name)
cur = conn.cursor()

db = create_engine("mysql+mysqlconnector://%s:%s@%s/%s" % (sqluser, sqlpassword, sqlhost, database_name))
db.echo = False
meta = MetaData()
meta.reflect(bind=db)
projects = Table('projects', meta, autoload=True)
con = db.connect()
logging.info("Connected to database")

attribute_list = ["Office", "Client", "Project number", "Sent to PM",
                  "Sent by PM", "PM First Name", "PM Last Name", "Date of upload"]
title_font = Font(bold=True)


# Constructs a set of all offices in the database
def build_office_set():
    stmt = select([projects.c.office])
    result = con.execute(stmt)
    office_set = set()
    for row in result:
        office_set.add(row[0])
    logging.info("Set of offices constructed from database")
    return office_set


def build_region_set():
    stmt = select([projects.c.region])
    result = con.execute(stmt)
    region_set = set()
    for row in result:
        region_set.add(row[0])
    logging.info("Set of regions constructed from database")
    return region_set


# Currently not in use
def get_pending(office):
    stmt = select([
        projects.c.office,
        projects.c.projectName,
        projects.c.pmName,
        projects.c.pmLastName,
        projects.c.subProjectNo,
        projects.c.customerName,
        projects.c.externalContact,
        projects.c.contact
    ]).where(and_(
        projects.c.office == office,
        projects.c.pmSendStatus == "no",
        projects.c.adminSendStatus == "yes"
    ))
    result = con.execute(stmt)

    out = []
    for row in result:
        out.append(row)
    return out


# Returns list of pending surveys for a given region
def get_pending_region(region):
    stmt = select([
        projects.c.office,
        projects.c.customerName,
        projects.c.subProjectNo,
        projects.c.adminSendStatus,
        projects.c.pmSendStatus,
        projects.c.pmName,
        projects.c.pmLastName,
        projects.c.dateUpload
    ]).where(and_(
        projects.c.region == region,
    )).order_by(
        projects.c.office,
        projects.c.subProjectNo,
        projects.c.adminSendStatus,
        projects.c.pmSendStatus)

    result = con.execute(stmt)
    logging.info("Queried database for surveys triggered")
    out = []
    for row in result:
        out.append(row)
    return out


def print_all_pending_by_office(offices):
    counter = 0
    for office in offices[:2]:
        out = get_pending(office)
        for s in out:
            print(s)
            counter += 1
        print("---")
    print("%i surveys missing" % counter)


def print_all_pending_by_region(regions):
    wb = Workbook()
    logging.info("Preparing results spreadsheet")
    for reg in regions:
        totals = {}
        ws = wb.create_sheet(title=reg)
        ws.cell(row=1, column=1).value = "Client Satisfaction Surveys for Region %s" % reg
        col = 1
        row = 3
        # Loop to write header row
        for header in attribute_list:
            ws.cell(row=row, column=col).value = header
            ws.cell(row=row, column=col).font = title_font
            col += 1
        row += 1
        col = 1

        # Collect list of pending surveys for the region
        missing_surveys = get_pending_region(reg)
        i = 0
        for survey in missing_surveys:
            totals[survey[0]] = []
            for val in survey:
                try:
                    totals[survey[0]][i] += 1
                except IndexError:
                    logging.info("Creating index for %s" % survey[0])
                    totals[survey[0]].append(0)
                i += 1
                ws.cell(row=row, column=col).value = val
                col += 1
                # print("¤ %s" % val)
            row += 1
            col = 1
            i = 0
            # print("---")

        row += 1
        ws.cell(row=row, column=col).value = "Totals by office"
        row += 1
        col = 1
        ws.cell(row=row, column=col).value = "Total sent by PM"
        col += 1
        ws.cell(row=row, column=col).value = "Total sent to client"
        col += 1
        ws.cell(row=row, column=col).value = "Total rejected"
        row += 1
        col = 1
        for key, values in totals.items():
            ws.cell(row=row, column=col).value = key
            col += 1
            ws.cell(row=row, column=col).value = values[0]
            col += 1
            ws.cell(row=row, column=col).value = values[1]
            col = 1
            row += 1

    # print("%i surveys missing" % counter)

    # Clean up autocreated blank sheets in workbook
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    logging.info("Cleaning up stuff...")
    # wb.remove_sheet(wb.get_sheet_by_name("Sheet1"))

    # Add a time/date stamp to filename and save
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d-%H%M")
    filename = "Missing CSS (regions) %s.xlsx" % timestamp
    wb.save(filename)
    logging.info("Excel file saved in program root with name %s" % filename)


def inspect_table(tablename):
    inspector = inspect(db)
    for column in inspector.get_columns(tablename):
        print(column['name'])
    # stmt = select([projects])
    # out = con.execute(stmt)
    # print(out.fetchall())


def get_fields():
    inspect_table("projects")


def main():
    # inspect_table("projects")
    offices = list(build_office_set())
    regions = list(build_region_set())

    offices.sort()
    regions.sort()

    print_all_pending_by_region(regions)
    logging.info("All done, shutting down. Enjoy.")

if __name__ == '__main__':
    main()